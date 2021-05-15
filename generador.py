#!/usr/bin/env python3
# Para crear el ejecutable:
# pyinstaller --onefile --name NOMBRE --noconsole --icon NOMBRE_ICONO generador.py
# formato_indice.xlsx y la carpeta de procesos deben estar en la misma carpeta que el ejecutable
import os, shutil
import PyPDF2
import tkinter
from os.path import isfile, join, isdir
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime


def contar_paginas_pdf(ruta_archivo):
    # Cargamos el documento PDF con permisos de lectura
    pdf_object = open(ruta_archivo, 'rb')
    # Abrimos el PDF con PyPDF2
    doc_pdf = PyPDF2.PdfFileReader(pdf_object)
    # Devolvemos el numero de paginas
    return doc_pdf.numPages


def extraer_fecha_creacion(ruta_archivo):
    try:
        pdf_object = open(ruta_archivo, 'rb')
        doc_pdf = PyPDF2.PdfFileReader(pdf_object)
        fecha_creacion_pdf = doc_pdf.documentInfo.get('/CreationDate')
        ano_creacion = fecha_creacion_pdf[2:6]
        mes_creacion = fecha_creacion_pdf[6:8]
        dia_creacion = fecha_creacion_pdf[8:10]
        fecha_creacion = dia_creacion + '/' + mes_creacion + '/' + ano_creacion
        return(fecha_creacion)
    except:
        return('sin fecha')


def calc_peso_carpeta(ruta_carpeta):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(ruta_carpeta):
        for file in filenames:
            filepath = join(dirpath, file)
            total_size += os.path.getsize(filepath)
    str_total_size = str(round(total_size/1024)) + ' KB'
    return str_total_size


def llenar_indice_electronico(ruta_carpeta_interna, contenido_carpeta_interna, ruta_indice):
    nombre_carpeta_interna = ruta_carpeta_interna.split("\\")[-1]
    # Ordenamos la lista por nombre
    contenido_carpeta_interna.sort()
    # Cargamos el indice electronico y seleccionamos la hoja
    wb = load_workbook(ruta_indice)
    sheet = wb.active
    # Llenado de Ciudad, Despacho, Serie Doc, Radicacion:
    sheet['B5'].value = ciudad
    sheet['B5'].alignment = Alignment(horizontal="left")
    sheet['B6'].value = despacho_judicial
    sheet['B6'].alignment = Alignment(horizontal="left")
    sheet['B8'].value = nombre_carpeta_interna
    sheet['B8'].alignment = Alignment(horizontal="left")
    sheet['J8'].value = num_carpetas
    # Creacion y llenado de filas:
    fila_inicial = 14
    thin_border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000')
                    )
    for index, arch_carpeta_interna in enumerate(contenido_carpeta_interna):
        ruta_archivo = join(ruta_carpeta_interna, arch_carpeta_interna)
        fila = fila_inicial + index
        extension_archivo = (arch_carpeta_interna.split(".")[-1])
        # Insertar fila
        sheet.insert_rows(fila)
        # Estilo de borde de celdas de fila
        sheet.cell(row=fila, column=1).border = thin_border
        sheet.cell(row=fila, column=2).border = thin_border
        sheet.cell(row=fila, column=3).border = thin_border
        sheet.cell(row=fila, column=4).border = thin_border
        sheet.cell(row=fila, column=5).border = thin_border
        sheet.cell(row=fila, column=6).border = thin_border
        sheet.cell(row=fila, column=7).border = thin_border
        sheet.cell(row=fila, column=8).border = thin_border
        sheet.cell(row=fila, column=9).border = thin_border
        sheet.cell(row=fila, column=10).border = thin_border
        sheet.cell(row=fila, column=11).border = thin_border
        # Nombre archivo
        sheet['A'+ str(fila)] = arch_carpeta_interna
        # Fecha creacion
        if extension_archivo == 'pdf':
            sheet['B'+ str(fila)] = extraer_fecha_creacion(ruta_archivo)
        else:
            fecha_archivo = datetime.fromtimestamp(os.path.getctime(ruta_archivo))
            sheet['B'+ str(fila)] = fecha_archivo.strftime('%d/%m/%Y')
        # Fecha Incorporacion
        sheet['C'+ str(fila)] = fecha_incorporacion
        # Orden
        sheet['D'+ str(fila)] = index + 1
        # Numero de paginas
        if extension_archivo == 'pdf':
            sheet['E'+ str(fila)] = contar_paginas_pdf(ruta_archivo)
            # Pagina inicio
            sheet['F'+ str(fila)] = 1
            # Pagina fin
            sheet['G'+ str(fila)] = sheet['E'+ str(fila)].value
        else:
            sheet['E'+ str(fila)] = '__'
            sheet['F'+ str(fila)] = '__'
            sheet['G'+ str(fila)] = '__'
        # Formato y Peso
        if os.path.isdir(ruta_archivo):
            # Formato carpetas
            sheet['H'+ str(fila)] = "CARPETA"
            # Peso carpetas
            sheet['I'+ str(fila)] = calc_peso_carpeta(ruta_archivo)
        else:
            # Formato archivos
            sheet['H'+ str(fila)] = extension_archivo.upper()
            # Peso archivos
            peso_archivo = str(round((os.path.getsize(ruta_archivo))/1024)) + ' KB'
            sheet['I'+ str(fila)] = peso_archivo
        # Origen
        sheet['J'+ str(fila)] = origen
        # Alineacion de Celdas
        sheet['B'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['C'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['D'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['E'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['F'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['G'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['H'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['I'+ str(fila)].alignment = Alignment(horizontal="center")
        sheet['J'+ str(fila)].alignment = Alignment(horizontal="center")
    # Guardamos el indice electronico con los cambios
    wb.save(ruta_indice)


def obtener_rutas_internas(ruta_carpeta_externa, ruta_indice):
    contenido_carpeta_externa = os.listdir(ruta_carpeta_externa)
    for arch_carpeta_externa in contenido_carpeta_externa:
        ruta_arch_carpeta_externa = join(ruta_carpeta_externa, arch_carpeta_externa)
        if isdir(ruta_arch_carpeta_externa):
            ruta_carpeta_interna = ruta_arch_carpeta_externa
            # Obtener archivos de la carpeta interna:
            contenido_carpeta_interna = os.listdir(ruta_carpeta_interna)
            # Llenamos el indice electronico:
            llenar_indice_electronico(ruta_carpeta_interna, contenido_carpeta_interna, ruta_indice)


def crear_indices_electronicos():
    try:
        contenido = os.listdir(ruta)
        # Crear array de rutas externas (ejm. archivos\\Proceso xxx )
        rutas_carpetas_externas = [join(ruta, nombre) for nombre in contenido if isdir(join(ruta, nombre))]
        # Recorremos el array y creamos los indices
        for ruta_carpeta_externa in rutas_carpetas_externas:
            nombre_carpeta_externa = ruta_carpeta_externa.split("\\")[1]
            nombre_indice = 'Índice Electrónico ' + nombre_carpeta_externa + '.xlsx'
            ruta_indice = join(ruta_carpeta_externa, nombre_indice)
            shutil.copy('formato_indice.xlsx', ruta_indice)
            # Obtener rutas de carpetas internas
            obtener_rutas_internas(ruta_carpeta_externa, ruta_indice)
        return('Índices Generados Correctamente')
    except:
        return('ERROR: en la generación de los archivos')


def capturar_datos_ventana():
    # Variables globales importantes
    global ruta
    global ciudad
    global despacho_judicial
    global num_carpetas
    global fecha_incorporacion
    global origen
    ruta = r'' + input_ruta.get() + ''
    ciudad = input_ciudad.get()
    despacho_judicial = input_despacho.get()
    num_carpetas = input_carpetas.get()
    fecha_incorporacion = input_incorporacion.get()
    origen = input_origen.get()
    # Iniciamos la logica:
    label_respuesta["text"] = crear_indices_electronicos()


def crear_ventana():
    app = tkinter.Tk()
    app.title("Generador de Índices Electrónicos")
    ventana = tkinter.Frame(app)
    ventana.grid(column=0, row=0, padx=(50, 50), pady=(10, 10))
    ventana.columnconfigure(0, weight=1)
    ventana.rowconfigure(0, weight=1)
    # Creamos elementos
    global input_ruta
    global input_ciudad
    global input_despacho
    global input_carpetas
    global input_incorporacion
    global input_origen
    global label_respuesta
    label_0_0 = tkinter.Label(ventana, height=2)
    label_ruta = tkinter.Label(ventana, text="Ruta")
    input_ruta = tkinter.Entry(ventana, width=25)
    input_ruta.insert(tkinter.END, 'procesos')
    label_3_0 = tkinter.Label(ventana, width=8)
    label_0_1 = tkinter.Label(ventana, height=2)
    label_ciudad = tkinter.Label(ventana, text="Ciudad")
    input_ciudad = tkinter.Entry(ventana, width=25)
    input_ciudad.insert(tkinter.END, 'TUMACO')
    label_0_2 = tkinter.Label(ventana, height=2)
    label_despacho = tkinter.Label(ventana, text="Despacho Judicial")
    input_despacho = tkinter.Entry(ventana, width=25)
    input_despacho.insert(tkinter.END, 'Juzgado Primero Administrativo del Circuito de Tumaco')
    label_0_3 = tkinter.Label(ventana, height=2)
    label_carpetas = tkinter.Label(ventana, text="Numero Carpetas o Legajos")
    input_carpetas = tkinter.Entry(ventana, width=25)
    input_carpetas.insert(tkinter.END, 'PENDIENTE')
    label_0_4 = tkinter.Label(ventana, height=2)
    label_incorporacion = tkinter.Label(ventana, text="Fecha de Incorporacion")
    input_incorporacion = tkinter.Entry(ventana, width=25)
    input_incorporacion.insert(tkinter.END, '00/00/0000')
    label_0_5 = tkinter.Label(ventana, height=2)
    label_origen = tkinter.Label(ventana, text="Origen")
    input_origen = tkinter.Entry(ventana, width=25)
    input_origen.insert(tkinter.END, 'DIGITALIZADO')
    btn_generar = tkinter.Button(ventana, text="Generar", command=capturar_datos_ventana)
    label_respuesta = tkinter.Label(ventana, font="Arial 13")
    # Posicionamos elementos
    label_0_0.grid(column=0, row=0)
    label_ruta.grid(column=1, row=0)
    input_ruta.grid(column=2, row=0)
    label_3_0.grid(column=3, row=0)
    label_0_1.grid(column=0, row=1)
    label_ciudad.grid(column=1, row=1)
    input_ciudad.grid(column=2, row=1)
    label_0_2.grid(column=0, row=2)
    label_despacho.grid(column=1, row=2)
    input_despacho.grid(column=2, row=2)
    label_0_3.grid(column=0, row=3)
    label_carpetas.grid(column=1, row=3)
    input_carpetas.grid(column=2, row=3)
    label_0_3.grid(column=0, row=4)
    label_incorporacion.grid(column=1, row=4)
    input_incorporacion.grid(column=2, row=4)
    label_0_3.grid(column=0, row=5)
    label_origen.grid(column=1, row=5)
    input_origen.grid(column=2, row=5)
    btn_generar.grid(column=1, row=6)
    label_respuesta.grid(column=2, row=6)
    ventana.mainloop()


if __name__ == '__main__':
    # Creamos la ventana Tkinter
    crear_ventana()